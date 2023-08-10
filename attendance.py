from openpyxl import load_workbook
import pandas as pd
import numpy as np
import datetime

filename = 'Attendance System.xlsx'

# Load name list
wb = load_workbook(filename=filename)
# Get list of all existing Excel sheets
sheet_list = wb.sheetnames

def read(employee_name):
    # Iterate through every sheet in Excel file
    for sheet in sheet_list:
        # Read individual Excel sheet
        ws = wb[sheet]
        excel_sheet = pd.read_excel(filename, sheet_name=sheet)

        # Check through every namea
        for names in excel_sheet['Name']:
            # Replace NaN values with empty values
            excel_sheet = excel_sheet.replace(np.nan, '')
            if employee_name == names:
                print('Found ' + employee_name)
                # Find row number of employee name
                details = excel_sheet.loc[excel_sheet['Name'] == employee_name]
                # Find Attendance column number
                attendance_col = excel_sheet.columns.get_loc('Attendance')
                # Find Timestamp column number
                timestamp_col = excel_sheet.columns.get_loc('Timestamp')
                # Display the time now
                now = datetime.datetime.now().strftime('%d-%b-%Y %H:%M:%S')
                # Mark attendance and timestamp
                ws.cell(row=(int(details.index.values)+2), column=(attendance_col+1), value='Yes')
                ws.cell(row=(int(details.index.values)+2), column=(timestamp_col+1), value=now)
                # Save Excel file
                wb.save(filename)
            else:
                print('Employee not found')
                continue

if __name__ == '__main__':
    while True:
        employee_name = str(input('Enter name: '))
        read(employee_name)
